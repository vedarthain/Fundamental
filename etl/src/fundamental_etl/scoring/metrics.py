"""Metrics computation — for each stock, compute the metrics its scorecard requires.

Inputs (per stock):
  - app.fundamentals_annual (ordered period_end ascending)
  - app.fundamentals_quarterly (ordered ascending)
  - app.screener_meta (current_price, market_cap_cr if populated)
  - golden.indicators.daily_signals (latest row)
  - golden.price_history (252-day window for relative returns + above-200ema share)

Outputs:
  - One row in app.metrics_snapshot per (symbol, snapshot_date), holding:
      - The numeric value for every formula referenced by the stock's scorecard
        (stored in `cluster_metrics` JSONB)
      - The maturity_tier
      - market_cap, current_price for context
      - score_status flag
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Optional

import psycopg
from openpyxl import load_workbook

from .formulas import REGISTRY as FORMULAS
from .scorecards import get_scorecard, get_scorecard_from, load_db_overrides


# ----------------------- Loaders -------------------------------------------

def load_annual(conn: psycopg.Connection, symbol: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute("""
            SELECT period_end, sales, expenses, operating_profit, other_income, depreciation,
                   interest, profit_before_tax, tax, net_profit, dividend_amount,
                   equity_share_capital, reserves, borrowings, other_liabilities,
                   total_liabilities, net_block, cwip, investments, other_assets,
                   total_assets, receivables, inventory, cash_and_bank, no_of_equity_shares,
                   cash_from_operating, cash_from_investing, cash_from_financing, net_cash_flow,
                   annual_close_price
            FROM app.fundamentals_annual
            WHERE symbol = %s
            ORDER BY period_end ASC
        """, (symbol,))
        return [{k: (float(v) if v is not None and not isinstance(v, (str, date)) else v)
                 for k, v in r.items()}
                for r in cur.fetchall()]


def load_quarterly(conn: psycopg.Connection, symbol: str) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute("""
            SELECT period_end, sales, expenses, other_income, depreciation, interest,
                   profit_before_tax, tax, net_profit, operating_profit
            FROM app.fundamentals_quarterly
            WHERE symbol = %s
            ORDER BY period_end ASC
        """, (symbol,))
        return [{k: (float(v) if v is not None and not isinstance(v, (str, date)) else v)
                 for k, v in r.items()}
                for r in cur.fetchall()]


def load_signals(conn_golden: psycopg.Connection, symbol: str) -> Optional[dict]:
    """Latest indicators row for a symbol (NSE symbol w/ .NS suffix in golden_db)."""
    g_symbol = f"{symbol}.NS"
    with conn_golden.cursor() as cur:
        cur.execute("""
            SELECT * FROM indicators.daily_signals
            WHERE symbol = %s
            ORDER BY date DESC
            LIMIT 1
        """, (g_symbol,))
        row = cur.fetchone()
        if not row:
            return None
        return {k: (float(v) if isinstance(v, (int,)) and k not in ('date',) else v if not hasattr(v, '__float__') else float(v))
                for k, v in row.items()}


def load_price_history(conn_golden: psycopg.Connection, symbol: str) -> list[dict]:
    """1-year of daily closes for the symbol.

    Filters close IS NOT NULL — golden.price_history can have rows where the
    daily ingest wrote the row but failed to populate close (e.g. partial
    yfinance fetch). Including those would make the most-recent row a NULL,
    breaking every relative-return formula downstream.
    """
    g_symbol = f"{symbol}.NS"
    with conn_golden.cursor() as cur:
        cur.execute("""
            SELECT date, close
            FROM golden.price_history
            WHERE symbol = %s AND interval = '1d' AND close IS NOT NULL
            ORDER BY date DESC
            LIMIT 260
        """, (g_symbol,))
        rows = cur.fetchall()
        return [{"date": r["date"], "close": float(r["close"])} for r in rows]


def load_above_200ema(conn_golden: psycopg.Connection, symbol: str) -> Optional[float]:
    """% of last 252 trading days above 200 EMA."""
    g_symbol = f"{symbol}.NS"
    with conn_golden.cursor() as cur:
        cur.execute("""
            SELECT
              SUM(CASE WHEN above_200ema THEN 1 ELSE 0 END)::float / NULLIF(COUNT(*), 0) AS pct
            FROM (
              SELECT above_200ema FROM indicators.daily_signals
              WHERE symbol = %s ORDER BY date DESC LIMIT 252
            ) x
        """, (g_symbol,))
        row = cur.fetchone()
        return float(row["pct"]) if row and row["pct"] is not None else None


def load_nifty_returns(conn_golden: psycopg.Connection) -> dict[str, float]:
    """Median NSE active-stock return over 3M / 6M / 12M.

    golden_db has no NIFTY index ticker; we compute a market-wide median benchmark
    instead. For relative-return scoring, only the location of the zero-line matters,
    and the median is more robust than the mean to outlier IPOs/penny stocks.
    """
    with conn_golden.cursor() as cur:
        # close IS NOT NULL throughout — a broken ingest day (rows present,
        # close blank) would otherwise corrupt every anchor date and median.
        cur.execute("""
            WITH latest AS (
              SELECT MAX(date) AS d FROM golden.price_history
               WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL
            ),
            windows AS (
              SELECT
                d AS today,
                d - INTERVAL '21 days'  AS d_1m,
                d - INTERVAL '63 days'  AS d_3m,
                d - INTERVAL '126 days' AS d_6m,
                d - INTERVAL '252 days' AS d_12m
              FROM latest
            ),
            base AS (
              SELECT ph.symbol, ph.date, ph.close
              FROM golden.price_history ph
              JOIN golden.stocks s USING (symbol)
              JOIN windows w ON TRUE
              WHERE ph.interval='1d' AND s.exchange='NSE' AND s.is_active
                AND ph.close IS NOT NULL
                AND ph.date IN (
                  -- nearest available trading day to each anchor
                  (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL),
                  (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_1m FROM windows)),
                  (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_3m FROM windows)),
                  (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_6m FROM windows)),
                  (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_12m FROM windows))
                )
            ),
            anchor_dates AS (
              SELECT
                (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL) AS d_now,
                (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_1m FROM windows)) AS d_1m,
                (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_3m FROM windows)) AS d_3m,
                (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_6m FROM windows)) AS d_6m,
                (SELECT MAX(date) FROM golden.price_history WHERE interval='1d' AND symbol LIKE '%.NS' AND close IS NOT NULL AND date <= (SELECT d_12m FROM windows)) AS d_12m
            ),
            now AS (SELECT symbol, close FROM base, anchor_dates WHERE base.date = anchor_dates.d_now),
            d1  AS (SELECT symbol, close FROM base, anchor_dates WHERE base.date = anchor_dates.d_1m),
            d3  AS (SELECT symbol, close FROM base, anchor_dates WHERE base.date = anchor_dates.d_3m),
            d6  AS (SELECT symbol, close FROM base, anchor_dates WHERE base.date = anchor_dates.d_6m),
            d12 AS (SELECT symbol, close FROM base, anchor_dates WHERE base.date = anchor_dates.d_12m)
            SELECT
              percentile_cont(0.5) WITHIN GROUP (ORDER BY now.close/d1.close - 1)  AS r1,
              percentile_cont(0.5) WITHIN GROUP (ORDER BY now.close/d3.close - 1)  AS r3,
              percentile_cont(0.5) WITHIN GROUP (ORDER BY now.close/d6.close - 1)  AS r6,
              percentile_cont(0.5) WITHIN GROUP (ORDER BY now.close/d12.close - 1) AS r12
            FROM now
              JOIN d1  USING (symbol)
              JOIN d3  USING (symbol)
              JOIN d6  USING (symbol)
              JOIN d12 USING (symbol)
            WHERE d1.close > 0 AND d3.close > 0 AND d6.close > 0 AND d12.close > 0
        """)
        r = cur.fetchone()
    if not r:
        return {}
    return {
        "1m":  float(r["r1"])  if r["r1"]  is not None else None,
        "3m":  float(r["r3"])  if r["r3"]  is not None else None,
        "6m":  float(r["r6"])  if r["r6"]  is not None else None,
        "12m": float(r["r12"]) if r["r12"] is not None else None,
    }


def compute_returns(prices: list[dict]) -> dict[str, Optional[float]]:
    """Compute 1M/3M/6M/12M returns from a daily price list (newest-first)."""
    if not prices:
        return {"ret_1m": None, "ret_3m": None, "ret_6m": None, "ret_12m": None}
    prices = list(reversed(prices))  # oldest first
    closes = [p["close"] for p in prices if p["close"] is not None]
    if not closes:
        return {"ret_1m": None, "ret_3m": None, "ret_6m": None, "ret_12m": None}
    last = closes[-1]

    def _ret(window):
        if len(closes) <= window:
            return None
        return last / closes[-window - 1] - 1

    return {
        "ret_1m": _ret(21),
        "ret_3m": _ret(63),
        "ret_6m": _ret(126),
        "ret_12m": _ret(252),
    }


# ----------------------- Meta from latest raw export ------------------------

def load_meta_from_raw(conn: psycopg.Connection, symbol: str) -> dict:
    """Read the most recent raw xlsx blob and pull mc / current_price / face_value etc."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT content FROM app.screener_export_raw
            WHERE symbol = %s ORDER BY fetched_at DESC LIMIT 1
        """, (symbol,))
        r = cur.fetchone()
        if not r:
            return {}
    wb = load_workbook(io.BytesIO(r["content"]), data_only=True, read_only=True)
    if "Data Sheet" not in wb.sheetnames:
        return {}
    ws = wb["Data Sheet"]
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row:
            continue
        label = (row[0] or "").strip() if isinstance(row[0], str) else None
        if label == "Current Price" and row[1] is not None:
            try: out["current_price"] = float(row[1])
            except: pass
        elif label == "Market Capitalization" and row[1] is not None:
            try: out["market_cap_cr"] = float(row[1])
            except: pass
        elif label == "Face Value" and row[1] is not None:
            try: out["face_value"] = float(row[1])
            except: pass
        elif label == "Number of shares" and row[1] is not None:
            try: out["no_of_shares"] = float(row[1])
            except: pass
    return out


# ----------------------- OI Spike Adjustment --------------------------------
#
# A large one-time "Other Income" entry in the latest quarter inflates
# net_profit (and therefore PBT) without reflecting recurring business
# performance.  Downstream metrics that consume net_profit — pe_ttm,
# np_cagr_*, roe_avg_above_threshold_*, np_growth_above_inflation_*,
# roce_* — all move in the stock's favour even though the underlying
# business didn't improve.
#
# We detect the spike and SUBTRACT the non-recurring excess from the latest
# quarterly row AND (when the quarter is the fiscal year-end) the latest
# annual row, BEFORE any formula runs.  The original DB rows are never
# modified — this is a local in-memory adjustment that only affects
# this scoring run.
#
# Thresholds are intentionally identical to those in web/src/lib/oi-alerts.ts
# so the UI warning and the score correction fire on exactly the same stocks.

_OI_PCT_PBT_THRESHOLD = 0.40   # other_income / pbt must exceed this
_OI_SPIKE_RATIO       = 5.0    # relative to 8-quarter prior average
_OI_MIN_CR            = 10.0   # minimum absolute value (₹ Cr) to care about
_OI_MIN_PRIOR_QTR     = 3      # need at least this many prior quarters for a baseline


def _oi_spike_adjustment(quarterly: list[dict]) -> tuple[float, float]:
    """Return (pbt_excess, np_excess) to subtract from the latest quarter.

    Both are zero when no spike is detected.

    pbt_excess = one-time other income above the recurring baseline (pre-tax)
    np_excess  = pbt_excess × (1 − effective_tax_rate)   (after-tax)
    """
    if len(quarterly) < 4:
        return 0.0, 0.0

    latest = quarterly[-1]
    oi  = latest.get("other_income")
    pbt = latest.get("profit_before_tax")
    np_ = latest.get("net_profit")

    if oi is None or pbt is None or np_ is None:
        return 0.0, 0.0
    if oi <= _OI_MIN_CR or pbt <= 0:
        return 0.0, 0.0
    if oi / pbt <= _OI_PCT_PBT_THRESHOLD:
        return 0.0, 0.0

    # Baseline: average of positive OI across prior 2–9 quarters
    prior_ois = [q.get("other_income") for q in quarterly[-9:-1]]
    prior_ois = [v for v in prior_ois if v is not None]
    if len(prior_ois) < _OI_MIN_PRIOR_QTR:
        return 0.0, 0.0

    avg_prior_oi = sum(max(v, 0.0) for v in prior_ois) / len(prior_ois)
    spike_ratio  = (oi / avg_prior_oi) if avg_prior_oi > 0 else float("inf")
    if spike_ratio <= _OI_SPIKE_RATIO:
        return 0.0, 0.0

    pbt_excess = oi - avg_prior_oi                          # pre-tax one-time amount
    tax_rate   = max(0.0, min(0.40, 1.0 - np_ / pbt))      # effective rate, clamped
    np_excess  = pbt_excess * (1.0 - tax_rate)              # after-tax
    return pbt_excess, np_excess


def _apply_oi_adjustment(
    quarterly: list[dict],
    annual: list[dict],
    pbt_excess: float,
    np_excess: float,
) -> tuple[list[dict], list[dict]]:
    """Return adjusted quarterly + annual lists (original lists never mutated).

    Latest quarterly row: pbt and net_profit are reduced by the excess.
    Latest annual row: same adjustment applied ONLY when annual period_end
    equals the quarterly period_end (i.e. the spike quarter IS the fiscal
    year-end so the inflated number already rolled into the annual total).
    """
    if pbt_excess == 0.0 and np_excess == 0.0:
        return quarterly, annual

    # --- quarterly ---
    q = {**quarterly[-1]}
    if q.get("profit_before_tax") is not None:
        q["profit_before_tax"] = q["profit_before_tax"] - pbt_excess
    if q.get("net_profit") is not None:
        q["net_profit"] = q["net_profit"] - np_excess
    adj_quarterly = quarterly[:-1] + [q]

    # --- annual (only when the fiscal year matches the spike quarter) ---
    adj_annual = annual
    if (
        annual
        and quarterly
        and annual[-1].get("period_end") == quarterly[-1].get("period_end")
    ):
        a = {**annual[-1]}
        if a.get("profit_before_tax") is not None:
            a["profit_before_tax"] = a["profit_before_tax"] - pbt_excess
        if a.get("net_profit") is not None:
            a["net_profit"] = a["net_profit"] - np_excess
        adj_annual = annual[:-1] + [a]

    return adj_quarterly, adj_annual


# ----------------------- Compute -------------------------------------------

def compute_metrics_for_symbol(
    app_conn: psycopg.Connection,
    golden_conn: psycopg.Connection,
    symbol: str,
    cluster_id: str,
    tier: str,
    nifty_returns: dict,
    scorecard_overrides: dict | None = None,
) -> tuple[dict, dict, str]:
    """Compute all metrics for one stock. Returns (cluster_metrics, meta, score_status)."""
    annual = load_annual(app_conn, symbol)
    quarterly = load_quarterly(app_conn, symbol)
    meta = load_meta_from_raw(app_conn, symbol)

    # Strip one-time Other Income spike from the latest quarter (and, when it
    # coincides with the fiscal year-end, from the latest annual row too).
    # This prevents a large non-recurring gain from inflating pe_ttm, np_cagr,
    # roe_avg_above_threshold, roce and other net-profit-derived metrics.
    pbt_excess, np_excess = _oi_spike_adjustment(quarterly)
    if pbt_excess > 0:
        quarterly, annual = _apply_oi_adjustment(quarterly, annual, pbt_excess, np_excess)

    # Update screener_meta with the freshly-extracted meta fields.
    #
    # IMPORTANT: current_price is NOT written here. NSE bhavcopy is the
    # single source of truth for LTP (refresh-ltp.py / GH Action writes it
    # daily from the official exchange data).  Screener's Excel export
    # carries an LTP too, but it can lag the bhavcopy by hours-to-days and
    # is an inferior source.  We deliberately skip it so the ETL never
    # introduces stale prices into production.
    #
    # market_cap_cr DOES come from Screener — NSE bhavcopy doesn't publish
    # market cap, so Screener is the practical source. A small staleness
    # window is acceptable here (mc moves slowly day-to-day).
    with app_conn.cursor() as cur:
        cur.execute("""
            UPDATE app.screener_meta
            SET market_cap_cr = %s,
                face_value = %s,
                no_of_shares = %s
            WHERE symbol = %s
        """, (meta.get("market_cap_cr"),
              meta.get("face_value"), meta.get("no_of_shares"), symbol))

    signals = load_signals(golden_conn, symbol)
    pct_above = load_above_200ema(golden_conn, symbol)
    prices = load_price_history(golden_conn, symbol)
    rets = compute_returns(prices)
    meta_full = {
        **meta,
        **rets,
        "pct_above_200ema_252d": pct_above,
    }

    sc = (get_scorecard_from(scorecard_overrides, cluster_id, tier)
          if scorecard_overrides is not None else get_scorecard(cluster_id, tier))
    needed_formulas = set(sc.quality) | set(sc.valuation) | set(sc.momentum)
    # Always compute the loss-maker fallbacks too (cheap)
    for fname, _share in sc.loss_maker_val_fallback:
        needed_formulas.add(fname)

    out: dict[str, Optional[float]] = {}
    for fname in needed_formulas:
        fn = FORMULAS.get(fname)
        if fn is None:
            out[fname] = None
            continue
        try:
            v = fn(annual, quarterly, meta_full, signals, nifty_returns)
            out[fname] = float(v) if v is not None else None
        except Exception:
            out[fname] = None

    # score_status: count what fraction of expected metrics came back null
    nonnull = sum(1 for v in out.values() if v is not None)
    if not needed_formulas:
        return out, meta_full, "no_scorecard"
    nonnull_share = nonnull / len(needed_formulas)
    if nonnull_share == 0:
        status = "insufficient_data"
    elif nonnull_share < 0.5:
        status = "partial-data"
    else:
        status = "full"
    return out, meta_full, status


def persist_metrics(
    conn: psycopg.Connection,
    symbol: str,
    snapshot_date: date,
    cluster_metrics: dict,
    meta: dict,
    tier: str,
    score_status: str,
) -> None:
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO app.metrics_snapshot
              (symbol, snapshot_date, market_cap, score_status, maturity_tier, cluster_metrics)
            VALUES (%s, %s, %s, %s, %s, %s::jsonb)
            ON CONFLICT (symbol, snapshot_date) DO UPDATE
              SET market_cap = EXCLUDED.market_cap,
                  score_status = EXCLUDED.score_status,
                  maturity_tier = EXCLUDED.maturity_tier,
                  cluster_metrics = EXCLUDED.cluster_metrics
        """, (
            symbol, snapshot_date, meta.get("market_cap_cr"), score_status, tier,
            psycopg.types.json.Json(cluster_metrics),
        ))
