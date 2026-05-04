"""Parse the `Data Sheet` tab of a Screener export xlsx into structured rows.

The Data Sheet is laid out as labelled rows. The first column is a label; subsequent
columns are values aligned to a "Report Date" header row that appears once per section.
Sections of interest (in order of appearance):
  - PROFIT & LOSS   (annual, ~10 years)
  - Quarters        (quarterly, ~10 quarters)
  - BALANCE SHEET   (annual)
  - CASH FLOW:      (annual)
  - PRICE:          (annual)
  - DERIVED:        (annual; mostly formulas — skip)

We extract the four sections that contain raw fundamentals and align them by period_end.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook


# Map Screener row labels → our column names. Anything not listed is ignored.
PNL_LABELS = {
    "Sales": "sales",
    "Expenses": "expenses",
    "Operating Profit": "operating_profit",
    "Other Income": "other_income",
    "Depreciation": "depreciation",
    "Interest": "interest",
    "Profit before tax": "profit_before_tax",
    "Tax": "tax",
    "Net profit": "net_profit",
    "Dividend Amount": "dividend_amount",
}

QUARTER_LABELS = {
    "Sales": "sales",
    "Expenses": "expenses",
    "Other Income": "other_income",
    "Depreciation": "depreciation",
    "Interest": "interest",
    "Profit before tax": "profit_before_tax",
    "Tax": "tax",
    "Net profit": "net_profit",
    "Operating Profit": "operating_profit",
}

BS_LABELS = {
    "Equity Share Capital": "equity_share_capital",
    "Reserves": "reserves",
    "Borrowings": "borrowings",
    "Other Liabilities": "other_liabilities",
    "Net Block": "net_block",
    "Capital Work in Progress": "cwip",
    "Investments": "investments",
    "Other Assets": "other_assets",
    "Receivables": "receivables",
    "Inventory": "inventory",
    "Cash & Bank": "cash_and_bank",
    "No. of Equity Shares": "no_of_equity_shares",
}
# BS has two "Total" rows (liab + assets) — handled specially in parsing.

CF_LABELS = {
    "Cash from Operating Activity": "cash_from_operating",
    "Cash from Investing Activity": "cash_from_investing",
    "Cash from Financing Activity": "cash_from_financing",
    "Net Cash Flow": "net_cash_flow",
}


@dataclass
class ParsedExport:
    company_name: Optional[str] = None
    face_value: Optional[float] = None
    current_price: Optional[float] = None
    market_cap: Optional[float] = None
    annual: dict[date, dict[str, float]] = field(default_factory=dict)
    quarterly: dict[date, dict[str, float]] = field(default_factory=dict)
    annual_close_price: dict[date, float] = field(default_factory=dict)


class ParseError(Exception):
    pass


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Sometimes Screener stores dates as strings
    try:
        return datetime.fromisoformat(str(v)).date()
    except ValueError:
        return None


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_export(xlsx_bytes: bytes) -> ParsedExport:
    """Parse the Data Sheet tab and return a structured ParsedExport."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=False)
    if "Data Sheet" not in wb.sheetnames:
        raise ParseError(f"'Data Sheet' tab missing; found {wb.sheetnames}")
    ws = wb["Data Sheet"]
    rows = list(ws.iter_rows(values_only=True))

    out = ParsedExport()

    # Section state machine: walk rows, switch section on section-header rows.
    section: Optional[str] = None
    period_dates: list[Optional[date]] = []
    bs_total_seen = 0  # BS has two "Total" rows; first = total liabilities, second = total assets

    for row in rows:
        if not row:
            continue
        label = (row[0] or "").strip() if isinstance(row[0], str) else row[0]
        if not isinstance(label, str):
            continue

        # Top-level metadata rows (before any section)
        if label == "COMPANY NAME":
            out.company_name = str(row[1]) if row[1] is not None else None
            continue
        if label == "Face Value":
            out.face_value = _to_float(row[1])
            continue
        if label == "Current Price":
            out.current_price = _to_float(row[1])
            continue
        if label == "Market Capitalization":
            out.market_cap = _to_float(row[1])
            continue

        # Section headers
        if label == "PROFIT & LOSS":
            section = "pnl"; period_dates = []
            continue
        if label == "Quarters":
            section = "quarters"; period_dates = []
            continue
        if label == "BALANCE SHEET":
            section = "bs"; period_dates = []; bs_total_seen = 0
            continue
        if label == "CASH FLOW:":
            section = "cf"; period_dates = []
            continue
        if label == "PRICE:":
            # PRICE row itself; values aligned to the most recent annual period_dates seen.
            # Screener emits PRICE *after* the cash-flow Report Date row, so use cf periods.
            for i, val in enumerate(row[1:], start=0):
                if i < len(period_dates) and period_dates[i] is not None:
                    p = _to_float(val)
                    if p is not None:
                        out.annual_close_price[period_dates[i]] = p
            section = None  # nothing else to consume
            continue
        if label == "DERIVED:":
            section = "skip"
            continue

        if section is None or section == "skip":
            continue

        # Within a section: "Report Date" is the period header for this section
        if label == "Report Date":
            period_dates = [_to_date(v) for v in row[1:]]
            continue

        if section == "pnl" and label in PNL_LABELS:
            col = PNL_LABELS[label]
            for i, val in enumerate(row[1:], start=0):
                if i >= len(period_dates) or period_dates[i] is None:
                    continue
                v = _to_float(val)
                if v is None:
                    continue
                out.annual.setdefault(period_dates[i], {})[col] = v

        elif section == "quarters" and label in QUARTER_LABELS:
            col = QUARTER_LABELS[label]
            for i, val in enumerate(row[1:], start=0):
                if i >= len(period_dates) or period_dates[i] is None:
                    continue
                v = _to_float(val)
                if v is None:
                    continue
                out.quarterly.setdefault(period_dates[i], {})[col] = v

        elif section == "bs":
            if label == "Total":
                # Two Total rows in BS: 1st = total liabilities, 2nd = total assets
                col = "total_liabilities" if bs_total_seen == 0 else "total_assets"
                bs_total_seen += 1
                for i, val in enumerate(row[1:], start=0):
                    if i >= len(period_dates) or period_dates[i] is None:
                        continue
                    v = _to_float(val)
                    if v is None:
                        continue
                    out.annual.setdefault(period_dates[i], {})[col] = v
            elif label in BS_LABELS:
                col = BS_LABELS[label]
                for i, val in enumerate(row[1:], start=0):
                    if i >= len(period_dates) or period_dates[i] is None:
                        continue
                    v = _to_float(val)
                    if v is None:
                        continue
                    out.annual.setdefault(period_dates[i], {})[col] = v

        elif section == "cf" and label in CF_LABELS:
            col = CF_LABELS[label]
            for i, val in enumerate(row[1:], start=0):
                if i >= len(period_dates) or period_dates[i] is None:
                    continue
                v = _to_float(val)
                if v is None:
                    continue
                out.annual.setdefault(period_dates[i], {})[col] = v

    # Merge annual_close_price into annual rows
    for d, p in out.annual_close_price.items():
        out.annual.setdefault(d, {})["annual_close_price"] = p

    return out
