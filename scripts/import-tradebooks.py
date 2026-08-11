#!/usr/bin/env python3
"""
Import multi-broker equity tradebooks into app.portfolio_transaction.

- Skips ETFs / mutual-fund units (implicit: a row is kept only if it resolves
  to a symbol in app.universe, the cash-equity master which has no ETFs).
- Unifies all brokers into one portfolio (user_id=1), deduped at trade level.
- Symbol resolution: ISIN -> exact symbol -> normalized company-name prefix.

Usage:
  python scripts/import-tradebooks.py            # dry run, prints summary
  python scripts/import-tradebooks.py --commit    # writes to DB
  DB_URL=... python scripts/import-tradebooks.py --commit   # target DB
"""
import os, sys, re, glob, shutil, hashlib
from datetime import datetime, date
import psycopg as psycopg2
import openpyxl
import csv

DOWNLOADS = os.path.expanduser("~/Downloads")
USER_ID = 1
COMMIT = "--commit" in sys.argv

def get_db_url():
    if os.environ.get("DB_URL"):
        return os.environ["DB_URL"]
    # default: prod Neon from etl/.env.local
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    envf = os.path.join(root, "etl", ".env.local")
    with open(envf) as f:
        for line in f:
            if line.startswith("NEON_APP_URL="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    raise RuntimeError("no DB_URL / NEON_APP_URL")

# ---------- reference universe ----------
def load_universe(cur):
    cur.execute("select symbol, isin, company_name from app.universe where is_active")
    by_isin, by_symbol, by_name = {}, {}, []
    for sym, isin, name in cur.fetchall():
        if isin:
            by_isin[isin.strip().upper()] = sym
        by_symbol[sym.upper()] = sym
        by_name.append((normalize_name(name), sym))
    return by_isin, by_symbol, by_name

def normalize_name(n):
    if not n:
        return ""
    n = n.upper()
    for junk in [" LIMITED", " LTD", " INDUSTRIES", " INDIA", ".", ",", "-", "&"]:
        n = n.replace(junk, " ")
    return re.sub(r"\s+", " ", n).strip()

def resolve(rec, ref):
    by_isin, by_symbol, by_name = ref
    # 1. ISIN
    if rec.get("isin"):
        s = by_isin.get(rec["isin"].strip().upper())
        if s:
            return s
    # 2. exact symbol
    if rec.get("raw_symbol"):
        s = by_symbol.get(rec["raw_symbol"].strip().upper())
        if s:
            return s
    # 3. normalized name prefix (unique)
    nn = normalize_name(rec.get("raw_name", ""))
    if len(nn) >= 4:
        hits = {sym for un, sym in by_name if un.startswith(nn) or nn.startswith(un)}
        if len(hits) == 1:
            return next(iter(hits))
    return None

# ---------- parsers: yield normalized dicts ----------
def num(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    return float(str(x).replace(",", "").strip())

def parse_zerodha(path):
    out = []
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            if not row.get("symbol"):
                continue
            out.append(dict(
                broker="zerodha", raw_symbol=row["symbol"].strip(),
                raw_name=row["symbol"].strip(), isin=(row.get("isin") or "").strip(),
                side=row["trade_type"].strip().lower(),
                quantity=num(row["quantity"]), price=num(row["price"]),
                trade_date=row["trade_date"].strip(),
                trade_time=(row.get("order_execution_time") or "").strip(),
                trade_id=(row.get("trade_id") or "").strip(),
                order_id=(row.get("order_id") or "").strip(),
                source_file=os.path.basename(path)))
    return out

def parse_fyers(path):
    out = []
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    # find header row (starts with "Name")
    hidx = next(i for i, r in enumerate(rows) if r and r[0] == "Name")
    hdr = rows[hidx]
    for r in rows[hidx + 1:]:
        if not r or not r[0] or r[0] == "":
            continue
        d = dict(zip(hdr, r))
        if d.get("Status", "").strip() != "Executed":
            continue
        name = d["Name"].strip()  # NSE:SYMBOL-EQ
        m = re.match(r"^[A-Z]+:(.+)-[A-Z]+$", name)
        sym = m.group(1) if m else name
        dt = d["Date & Time"].strip()  # 05-08-2026 09:28:38
        dd, tt = (dt.split(" ", 1) + [""])[:2]
        out.append(dict(
            broker="fyers", raw_symbol=sym, raw_name=sym, isin="",
            side=d["Side"].strip().lower(),
            quantity=num(d["Qty"]), price=num(d["Traded price"]),
            trade_date=to_iso(dd), trade_time=tt,
            trade_id=re.sub(r'[^0-9.]', '', d.get("Exchange order ID", "")),
            order_id=re.sub(r'[^0-9.]', '', d.get("OMS order ID", "")),
            source_file=os.path.basename(path)))
    return out

def _xlsx_rows(path, sheet=None):
    if path.lower().endswith(".xls"):
        dst = "/tmp/_tb_" + os.path.basename(path) + ".xlsx"
        shutil.copy(path, dst)
        path = dst
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]

def _fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def to_iso(s):
    """Normalize a date token to YYYY-MM-DD. Accepts ISO or DD-MM-YYYY."""
    s = str(s).strip().split(" ")[0]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return s

def parse_groww(path):
    rows = _xlsx_rows(path)
    hidx = next(i for i, r in enumerate(rows)
                if r and str(r[0]).strip() == "Stock name")
    hdr = [str(c).strip() if c else "" for c in rows[hidx]]
    out = []
    for r in rows[hidx + 1:]:
        d = dict(zip(hdr, r))
        if not d.get("Symbol"):
            continue
        if str(d.get("Order status", "")).strip() != "Executed":
            continue
        out.append(dict(
            broker="groww", raw_symbol=str(d["Symbol"]).strip(),
            raw_name=str(d.get("Stock name", "")).strip(),
            isin=str(d.get("ISIN", "")).strip(),
            side=str(d["Type"]).strip().lower(),
            quantity=num(d["Quantity"]),
            price=(num(d["Value"]) / num(d["Quantity"])) if d.get("Value") and num(d["Quantity"]) else None,
            trade_date=to_iso(_fmt_date(d.get("Execution date and time"))),
            trade_time=str(d.get("Execution date and time", "")).strip(),
            trade_id="", order_id=str(d.get("Exchange Order Id", "")).strip(),
            source_file=os.path.basename(path)))
    return out

def parse_upstox(path):
    rows = _xlsx_rows(path, "TRADE")
    hidx = next(i for i, r in enumerate(rows)
                if r and str(r[0]).strip() == "Date")
    hdr = [str(c).strip() if c else "" for c in rows[hidx]]
    out = []
    for r in rows[hidx + 1:]:
        d = dict(zip(hdr, r))
        if not d.get("Company"):
            continue
        out.append(dict(
            broker="upstox", raw_symbol=str(d.get("Scrip Code", "")).strip(),
            raw_name=str(d["Company"]).strip(), isin="",
            side=str(d["Side"]).strip().lower(),
            quantity=num(d["Quantity"]), price=num(d["Price"]),
            trade_date=_fmt_date(d.get("Date")),
            trade_time=str(d.get("Trade Time", "")).strip(),
            trade_id=str(d.get("Trade Num", "")).strip(), order_id="",
            source_file=os.path.basename(path)))
    return out

def parse_5paisa(path):
    rows = _xlsx_rows(path)
    hidx = next((i for i, r in enumerate(rows)
                 if r and str(r[0]).strip() == "Transaction Date"), None)
    if hidx is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hidx]]
    out = []
    for r in rows[hidx + 1:]:
        d = dict(zip(hdr, r))
        if not d.get("Company Name"):
            continue
        out.append(dict(
            broker="fivepaisa", raw_symbol="",
            raw_name=str(d["Company Name"]).strip(), isin="",
            side=str(d["Type"]).strip().lower(),
            quantity=num(d["Quantity"]), price=num(d["Price"]),
            trade_date=_fmt_date(d.get("Transaction Date")),
            trade_time="", trade_id="", order_id="",
            source_file=os.path.basename(path)))
    return out

def collect():
    recs = []
    for p in glob.glob(f"{DOWNLOADS}/tradebook-*-EQ*.csv"):
        recs += parse_zerodha(p)
    for p in glob.glob(f"{DOWNLOADS}/FYERS_orderbook_*.csv"):
        recs += parse_fyers(p)
    for p in glob.glob(f"{DOWNLOADS}/Stocks_Order_History_*.xlsx"):
        recs += parse_groww(p)
    for p in glob.glob(f"{DOWNLOADS}/trade_*.xlsx"):
        recs += parse_upstox(p)
    for p in glob.glob(f"{DOWNLOADS}/Trade_Report_Equity_*.xls"):
        recs += parse_5paisa(p)
    return recs

def dedup_key(r):
    if r.get("trade_id"):
        raw = f"{r['broker']}|tid|{r['trade_id']}"
    else:
        raw = f"{r['broker']}|{r['trade_date']}|{r['symbol']}|{r['side']}|{r['quantity']}|{r['price']}|{r.get('trade_time','')}"
    return hashlib.md5(raw.encode()).hexdigest()

def recompute_derived_holding(cur, user_id, symbol):
    """Rebuild the synthetic broker='derived' portfolio_holding row for one
    (user, symbol) from that user's transactions. 1:1 mirror of
    web/src/lib/derivedHoldings.ts::recomputeDerivedHolding — average-cost walk,
    snapshot-wins, imported_at = first trade date. Idempotent."""
    # Snapshot wins: a real broker already reports this symbol -> drop any
    # derived row and stop. (Real brokers = everything except 'derived'.)
    cur.execute(
        "select 1 from app.portfolio_holding "
        "where user_id=%s and broker<>'derived' and symbol=%s limit 1",
        (user_id, symbol))
    if cur.fetchone():
        cur.execute(
            "delete from app.portfolio_holding "
            "where user_id=%s and broker='derived' and raw_symbol=%s",
            (user_id, symbol))
        return

    cur.execute(
        "select side, trade_date::text, quantity::float8, price::float8 "
        "from app.portfolio_transaction where user_id=%s and symbol=%s "
        "order by trade_date asc, trade_time asc nulls first, id asc",
        (user_id, symbol))
    qty = 0.0
    avg = 0.0
    first_date = None
    for side, d, q, price in cur.fetchall():
        if first_date is None:
            first_date = d
        if side == "buy":
            nxt = qty + q
            avg = (qty * avg + q * price) / nxt if nxt > 0 else 0.0
            qty = nxt
        else:
            qty -= q
    qty = round(qty, 4)

    if qty <= 0:
        cur.execute(
            "delete from app.portfolio_holding "
            "where user_id=%s and broker='derived' and raw_symbol=%s",
            (user_id, symbol))
        return

    cur.execute("select isin from app.universe where symbol=%s limit 1", (symbol,))
    row = cur.fetchone()
    isin = row[0] if row else None
    avg_cost = round(avg, 4) if avg > 0 else None
    imported_at = first_date or date.today().isoformat()

    cur.execute("""
        insert into app.portfolio_holding
          (user_id, broker, raw_symbol, isin, symbol, is_mapped, quantity,
           avg_cost, source_batch, imported_at)
        values (%s,'derived',%s,%s,%s,true,%s,%s,gen_random_uuid(),%s)
        on conflict (user_id, broker, raw_symbol) do update
          set quantity=excluded.quantity,
              avg_cost=excluded.avg_cost,
              isin=excluded.isin,
              symbol=excluded.symbol,
              is_mapped=true,
              imported_at=excluded.imported_at
    """, (user_id, symbol, isin, symbol, qty, avg_cost, imported_at))

DDL = """
create table if not exists app.portfolio_transaction (
    id           bigserial primary key,
    user_id      bigint not null,
    broker       text   not null,
    trade_date   date   not null,
    trade_time   text,
    side         text   not null check (side in ('buy','sell')),
    symbol       text   not null,
    raw_symbol   text,
    raw_name     text,
    isin         text,
    quantity     numeric not null,
    price        numeric not null,
    trade_id     text,
    order_id     text,
    source_file  text,
    dedup_key    text not null unique,
    imported_at  timestamptz not null default now()
);
create index if not exists idx_ptx_user_symbol on app.portfolio_transaction(user_id, symbol);
create index if not exists idx_ptx_symbol_date on app.portfolio_transaction(symbol, trade_date);
"""

def main():
    db = get_db_url()
    conn = psycopg2.connect(db)
    cur = conn.cursor()
    ref = load_universe(cur)
    recs = collect()

    kept, dropped = [], {}
    seen = set()
    for r in recs:
        if r["side"] not in ("buy", "sell"):
            dropped.setdefault(f"bad-side:{r['broker']}", 0)
            dropped[f"bad-side:{r['broker']}"] += 1
            continue
        sym = resolve(r, ref)
        if not sym:
            dropped.setdefault(f"etf/unmapped:{r['broker']}", 0)
            dropped[f"etf/unmapped:{r['broker']}"] += 1
            continue
        r["symbol"] = sym
        k = dedup_key(r)
        if k in seen:
            dropped.setdefault("intra-run-dup", 0)
            dropped["intra-run-dup"] += 1
            continue
        seen.add(k)
        r["dedup_key"] = k
        kept.append(r)

    # summary
    print(f"\nDB: {db.split('@')[-1].split('/')[0]}")
    print(f"parsed {len(recs)} raw rows -> keeping {len(kept)} equity trades")
    from collections import Counter
    bc = Counter(r["broker"] for r in kept)
    for b in sorted(bc):
        buys = sum(1 for r in kept if r["broker"] == b and r["side"] == "buy")
        sells = sum(1 for r in kept if r["broker"] == b and r["side"] == "sell")
        print(f"  {b:10s}: {bc[b]:3d}  (buy {buys}, sell {sells})")
    print("dropped:")
    for k in sorted(dropped):
        print(f"  {k}: {dropped[k]}")
    syms = sorted(set(r["symbol"] for r in kept))
    print(f"distinct symbols: {len(syms)}")
    dr = sorted(r["trade_date"] for r in kept)
    print(f"date range: {dr[0]} -> {dr[-1]}")

    if not COMMIT:
        print("\nDRY RUN — pass --commit to write.")
        return

    cur.execute(DDL)

    # CSV takes precedence over hand entry: drop any manual entry that matches an
    # imported trade on (symbol, broker, trade_date, quantity) — it's the same
    # real trade typed in by hand, so keeping both would double-count. Mirror of
    # web/src/app/api/portfolio/import-trades/route.ts. Dedup tuple is the product
    # spec's, NOT dedup_key (which also keys on side/price/time).
    superseded = 0
    seen_tuples = set()
    for r in kept:
        key = (r["symbol"], r["broker"], r["trade_date"], r["quantity"])
        if key in seen_tuples:
            continue
        seen_tuples.add(key)
        cur.execute("""
            delete from app.portfolio_transaction
             where user_id=%s and source_file='manual-entry'
               and broker=%s and symbol=%s and trade_date=%s and quantity=%s
        """, (USER_ID, r["broker"], r["symbol"], r["trade_date"], r["quantity"]))
        superseded += cur.rowcount

    ins = 0
    for r in kept:
        cur.execute("""
            insert into app.portfolio_transaction
              (user_id,broker,trade_date,trade_time,side,symbol,raw_symbol,raw_name,
               isin,quantity,price,trade_id,order_id,source_file,dedup_key)
            values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            on conflict (dedup_key) do nothing
        """, (USER_ID, r["broker"], r["trade_date"], r.get("trade_time"),
              r["side"], r["symbol"], r.get("raw_symbol"), r.get("raw_name"),
              r.get("isin") or None, r["quantity"], r["price"],
              r.get("trade_id") or None, r.get("order_id") or None,
              r["source_file"], r["dedup_key"]))
        ins += cur.rowcount

    # Recompute the transaction-derived holdings for every symbol this import
    # touched. A symbol with a real broker snapshot keeps the snapshot
    # (snapshot-wins) and its derived row is dropped; a symbol without one gets a
    # synthetic broker='derived' position computed from its trades. 1:1 mirror of
    # web/src/lib/derivedHoldings.ts, run in the same transaction as the inserts.
    symbols = sorted({r["symbol"] for r in kept if r.get("symbol")})
    for sym in symbols:
        recompute_derived_holding(cur, USER_ID, sym)

    conn.commit()
    print(f"\nCOMMITTED: inserted {ins} new rows (dedup_key conflicts skipped).")
    if superseded:
        print(f"superseded {superseded} manual entries matched by CSV trades.")
    if symbols:
        print(f"recomputed derived holdings for {len(symbols)} symbols.")
    cur.execute("select count(*) from app.portfolio_transaction where user_id=%s", (USER_ID,))
    print(f"table now holds {cur.fetchone()[0]} rows for user {USER_ID}.")

if __name__ == "__main__":
    main()
